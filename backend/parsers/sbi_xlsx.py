"""SBI XLSX (encrypted) bank statement parser.

SBI provides password-protected .xlsx files. Uses msoffcrypto to decrypt,
then openpyxl to read. The header row is found dynamically (typically around
row 17-18) and data rows follow until a "Statement Summary" sentinel.
"""

import io
import tempfile
from typing import Optional

import msoffcrypto
import openpyxl
import math

from .base import BankStatementParser, RawTransaction


class SBIXlsxParser(BankStatementParser):
    bank_name = "SBI"

    # Expected column headers (lowered) in the data table
    HEADER_KEYWORDS = ["date", "details", "ref no", "cheque no", "debit", "credit", "balance"]

    def can_parse(self, text: str, tables: list) -> bool:
        """For XLSX files, text/tables come from the detector's peek logic."""
        text_lower = text.lower()
        if "state bank" in text_lower or "sbi" in text_lower:
            return True
        matched = sum(1 for kw in self.HEADER_KEYWORDS if kw in text_lower)
        return matched >= 3

    def parse(self, file_path: str, password: Optional[str] = None) -> list[RawTransaction]:
        wb = self._open_workbook(file_path, password)
        ws = wb.active
        if ws is None:
            return []

        rows = list(ws.iter_rows(values_only=True))
        header_idx = self._find_header_row(rows)
        if header_idx is None:
            return []

        col_map = self._map_columns(rows[header_idx])
        if not col_map:
            return []

        transactions: list[RawTransaction] = []
        for row in rows[header_idx + 1:]:
            # Stop at summary section
            first_cell = str(row[0]) if row[0] is not None else ""
            if "statement summary" in first_cell.lower():
                break

            tx = self._parse_row(row, col_map)
            if tx:
                transactions.append(tx)

        return transactions

    # ------------------------------------------------------------------
    # Internal helpers
    # ------------------------------------------------------------------

    def _open_workbook(self, file_path: str, password: Optional[str] = None) -> openpyxl.Workbook:
        """Open workbook, decrypting with msoffcrypto if a password is given."""
        if password:
            with open(file_path, "rb") as f:
                office_file = msoffcrypto.OfficeFile(f)
                office_file.load_key(password=password)
                decrypted = io.BytesIO()
                office_file.decrypt(decrypted)
                decrypted.seek(0)
                return openpyxl.load_workbook(decrypted, data_only=True)
        else:
            return openpyxl.load_workbook(file_path, data_only=True)

    def _find_header_row(self, rows: list[tuple]) -> Optional[int]:
        """Find the row index containing column headers like Date, Details, Debit, Credit."""
        for i, row in enumerate(rows):
            row_text = " ".join(str(c).lower() for c in row if c is not None)
            if "date" in row_text and "debit" in row_text and "credit" in row_text:
                return i
        return None

    def _map_columns(self, header_row: tuple) -> Optional[dict]:
        col_map: dict[str, int] = {}
        for i, cell in enumerate(header_row):
            if cell is None:
                continue
            h = str(cell).strip().lower()
            if "date" in h and "date" not in col_map:
                col_map["date"] = i
            elif "detail" in h or "description" in h or "narration" in h or "particular" in h:
                col_map["narration"] = i
            elif "ref" in h or "cheque" in h:
                col_map["reference"] = i
            elif "debit" in h or "withdrawal" in h:
                col_map["debit"] = i
            elif "credit" in h or "deposit" in h:
                col_map["credit"] = i
            elif "balance" in h:
                col_map["balance"] = i

        if "date" in col_map and "narration" in col_map and ("debit" in col_map or "credit" in col_map):
            return col_map
        return None

    def _is_nan(self, value) -> bool:
        """Check if a value is NaN or None."""
        if value is None:
            return True
        if isinstance(value, float) and math.isnan(value):
            return True
        return False

    def _clean_narration(self, value) -> str:
        """Clean narration: replace embedded newlines with spaces, strip."""
        if value is None:
            return ""
        text = str(value).replace("\n", " ").replace("\r", " ")
        # Collapse multiple spaces
        return " ".join(text.split())

    def _parse_row(self, row: tuple, col_map: dict) -> Optional[RawTransaction]:
        try:
            date_val = row[col_map["date"]]
            if self._is_nan(date_val):
                return None

            # Parse date
            parsed_date = self._parse_date(str(date_val).strip(), [])
            if not parsed_date:
                return None

            narration = self._clean_narration(row[col_map["narration"]])
            if not narration:
                return None

            # Skip summary rows
            if "statement summary" in narration.lower():
                return None

            debit = self._parse_cell_amount(row[col_map["debit"]]) if "debit" in col_map else None
            credit = self._parse_cell_amount(row[col_map["credit"]]) if "credit" in col_map else None
            balance = self._parse_cell_amount(row[col_map["balance"]]) if "balance" in col_map else None
            reference = None
            if "reference" in col_map:
                ref_val = row[col_map["reference"]]
                if ref_val is not None and not self._is_nan(ref_val):
                    reference = str(ref_val).strip() or None

            if credit and credit > 0:
                return RawTransaction(
                    date=parsed_date, narration=narration, amount=credit,
                    tx_type="credit", balance=balance, reference=reference,
                )
            elif debit and debit > 0:
                return RawTransaction(
                    date=parsed_date, narration=narration, amount=debit,
                    tx_type="debit", balance=balance, reference=reference,
                )
            return None
        except (IndexError, TypeError, ValueError):
            return None

    def _parse_cell_amount(self, value) -> Optional[float]:
        """Parse amount from a cell that may be numeric or string."""
        if value is None:
            return None
        if isinstance(value, (int, float)):
            if math.isnan(value):
                return None
            return round(float(value), 2) if value else None
        # Fall back to string parsing
        return self._parse_amount(str(value))
