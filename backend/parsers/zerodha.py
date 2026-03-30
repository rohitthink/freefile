"""Parser for Zerodha Tax P&L XLSX reports.

Parses the 'Tradewise Exits' sheet for individual trades (with section headers
like 'Equity - Intraday', 'Equity - Short Term', etc.) and summary sheets
for aggregate P&L. Also parses 'Equity Dividends' sheet.
"""

import openpyxl
from dataclasses import dataclass, field
from datetime import datetime, date
from typing import Optional


@dataclass
class TradingEntry:
    """A single realized trade entry."""
    source: str = ""
    scrip_name: str = ""
    isin: str = ""
    asset_type: str = ""
    buy_date: Optional[str] = None
    sell_date: Optional[str] = None
    quantity: float = 0.0
    buy_value: float = 0.0
    sell_value: float = 0.0
    expenses: float = 0.0
    gain_loss: float = 0.0
    gain_type: str = ""
    holding_period_days: int = 0


@dataclass
class DividendEntry:
    """A single dividend record."""
    source: str = ""
    scrip_name: str = ""
    isin: str = ""
    ex_date: Optional[str] = None
    amount: float = 0.0
    tds: float = 0.0


@dataclass
class ZerodhaParseResult:
    """Complete result from parsing Zerodha P&L."""
    trades: list = field(default_factory=list)
    dividends: list = field(default_factory=list)
    summary: dict = field(default_factory=dict)


def _safe_float(val) -> float:
    if val is None:
        return 0.0
    try:
        f = float(val)
        return round(f, 2) if f == f else 0.0  # NaN check
    except (ValueError, TypeError):
        return 0.0


def _safe_str(val) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _parse_date_flexible(val) -> Optional[str]:
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    if not s:
        return None
    for fmt in ["%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%d-%b-%Y", "%d %b %Y"]:
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


# Section header → (asset_type, gain_type)
SECTION_MAP = {
    "equity - intraday": ("equity_intraday", "speculative"),
    "equity - short term": ("equity", "STCG_111A"),
    "equity - long term": ("equity", "LTCG_112A"),
    "equity - buyback": ("equity", "STCG_111A"),
    "non equity": ("non_equity", "STCG_111A"),
    "mutual funds": ("mutual_fund", "STCG_111A"),
    "f&o": ("fno", "fno_business"),
    "currency": ("currency", "fno_business"),
    "commodity": ("commodity", "fno_business"),
}


def _map_header_cols(ws, header_row: int) -> dict[str, int]:
    """Map column names to column indices (1-based) from a header row."""
    col_map = {}
    for c in range(1, ws.max_column + 1):
        h = _safe_str(ws.cell(header_row, c).value).lower()
        if "symbol" in h or "scrip" in h:
            col_map["symbol"] = c
        elif h == "isin":
            col_map["isin"] = c
        elif "entry date" in h or "buy date" in h:
            col_map["buy_date"] = c
        elif "exit date" in h or "sell date" in h:
            col_map["sell_date"] = c
        elif "quantity" in h or "qty" in h:
            col_map["quantity"] = c
        elif "buy value" in h:
            col_map["buy_value"] = c
        elif "sell value" in h:
            col_map["sell_value"] = c
        elif h == "profit" or "p&l" in h or "taxable profit" in h:
            if "pnl" not in col_map:
                col_map["pnl"] = c
        elif "period of holding" in h:
            col_map["holding"] = c
    return col_map


def parse_zerodha_pnl(file_path: str) -> ZerodhaParseResult:
    """Parse Zerodha Tax P&L XLSX file."""
    result = ZerodhaParseResult()
    wb = openpyxl.load_workbook(file_path, data_only=True)

    # Find the Tradewise Exits sheet
    tradewise_sheet = None
    equity_summary_sheet = None
    fno_summary_sheet = None
    dividend_sheet = None

    for name in wb.sheetnames:
        nl = name.lower()
        if "tradewise" in nl or "exits" in nl:
            tradewise_sheet = name
        elif "equity" in nl and "dividend" not in nl and nl != "non equity":
            equity_summary_sheet = name
        elif "f&o" in nl or "fno" in nl:
            fno_summary_sheet = name
        elif "dividend" in nl:
            dividend_sheet = name

    # --- Parse Tradewise Exits sheet (main trade data) ---
    if tradewise_sheet:
        ws = wb[tradewise_sheet]
        current_section = None
        current_asset_type = "equity"
        current_gain_type = "STCG_111A"
        current_col_map = {}

        for r in range(1, ws.max_row + 1):
            # Read row values
            vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            non_none = [v for v in vals if v is not None and _safe_str(v)]

            # Check for section header (single non-empty cell that matches known sections)
            if len(non_none) == 1 and isinstance(non_none[0], str):
                section_text = non_none[0].strip().lower()
                for key, (asset_type, gain_type) in SECTION_MAP.items():
                    if key in section_text:
                        current_section = key
                        current_asset_type = asset_type
                        current_gain_type = gain_type
                        current_col_map = {}
                        break
                continue

            # Check for table header row (has "Symbol" column)
            first_str_vals = [_safe_str(v).lower() for v in vals[:5] if v is not None]
            if "symbol" in first_str_vals:
                current_col_map = _map_header_cols(ws, r)
                continue

            # Parse trade data row
            if not current_col_map or "symbol" not in current_col_map:
                continue

            symbol = _safe_str(ws.cell(r, current_col_map["symbol"]).value)
            if not symbol:
                continue
            # Stop at total/summary rows
            if any(kw in symbol.lower() for kw in ["total", "grand", "profit", "loss"]):
                continue

            isin = _safe_str(ws.cell(r, current_col_map["isin"]).value) if "isin" in current_col_map else ""
            buy_date = _parse_date_flexible(ws.cell(r, current_col_map["buy_date"]).value) if "buy_date" in current_col_map else None
            sell_date = _parse_date_flexible(ws.cell(r, current_col_map["sell_date"]).value) if "sell_date" in current_col_map else None
            quantity = _safe_float(ws.cell(r, current_col_map["quantity"]).value) if "quantity" in current_col_map else 0.0
            buy_value = _safe_float(ws.cell(r, current_col_map["buy_value"]).value) if "buy_value" in current_col_map else 0.0
            sell_value = _safe_float(ws.cell(r, current_col_map["sell_value"]).value) if "sell_value" in current_col_map else 0.0
            pnl = _safe_float(ws.cell(r, current_col_map["pnl"]).value) if "pnl" in current_col_map else sell_value - buy_value
            holding_days = int(_safe_float(ws.cell(r, current_col_map["holding"]).value)) if "holding" in current_col_map else 0

            # For equity, classify based on actual holding period
            gain_type = current_gain_type
            if current_asset_type == "equity" and buy_date and sell_date:
                try:
                    bd = datetime.strptime(buy_date, "%Y-%m-%d").date()
                    sd = datetime.strptime(sell_date, "%Y-%m-%d").date()
                    days = (sd - bd).days
                    gain_type = "LTCG_112A" if days > 365 else "STCG_111A"
                except (ValueError, TypeError):
                    pass

            result.trades.append(TradingEntry(
                source="zerodha",
                scrip_name=symbol,
                isin=isin,
                asset_type=current_asset_type,
                buy_date=buy_date,
                sell_date=sell_date,
                quantity=quantity,
                buy_value=buy_value,
                sell_value=sell_value,
                expenses=0.0,
                gain_loss=pnl,
                gain_type=gain_type,
                holding_period_days=holding_days,
            ))

    # --- Parse summary from Equity sheet ---
    if equity_summary_sheet:
        ws = wb[equity_summary_sheet]
        for r in range(1, min(25, ws.max_row + 1)):
            label = _safe_str(ws.cell(r, 2).value).lower()
            val = _safe_float(ws.cell(r, 3).value)
            if "intraday" in label and "profit" in label:
                result.summary["intraday_profit"] = val
            elif "short term" in label and "profit" in label:
                result.summary["short_term_profit"] = val
            elif "long term" in label and "profit" in label:
                result.summary["long_term_profit"] = val
            elif "non equity" in label and "profit" in label:
                result.summary["non_equity_profit"] = val

    # --- Parse Equity Dividends sheet ---
    if dividend_sheet:
        ws = wb[dividend_sheet]
        col_map = {}
        for r in range(1, ws.max_row + 1):
            first_val = _safe_str(ws.cell(r, 2).value).lower()

            # Find header row
            if "symbol" in first_val:
                for c in range(1, ws.max_column + 1):
                    h = _safe_str(ws.cell(r, c).value).lower()
                    if "symbol" in h:
                        col_map["symbol"] = c
                    elif "isin" in h:
                        col_map["isin"] = c
                    elif "ex-date" in h or "ex date" in h or "date" in h:
                        col_map["ex_date"] = c
                    elif "dividend per share" in h or "dps" in h:
                        col_map["dps"] = c
                    elif "quantity" in h or "qty" in h:
                        col_map["quantity"] = c
                    elif "net" in h and "amount" in h:
                        col_map["amount"] = c
                continue

            if not col_map or "symbol" not in col_map:
                continue

            symbol = _safe_str(ws.cell(r, col_map["symbol"]).value)
            if not symbol or "total" in symbol.lower() or "grand" in symbol.lower():
                continue

            isin = _safe_str(ws.cell(r, col_map["isin"]).value) if "isin" in col_map else ""
            ex_date = _parse_date_flexible(ws.cell(r, col_map["ex_date"]).value) if "ex_date" in col_map else None

            amount = 0.0
            if "amount" in col_map:
                amount = _safe_float(ws.cell(r, col_map["amount"]).value)
            elif "dps" in col_map and "quantity" in col_map:
                dps = _safe_float(ws.cell(r, col_map["dps"]).value)
                qty = _safe_float(ws.cell(r, col_map["quantity"]).value)
                amount = round(dps * qty, 2)

            if amount != 0.0:
                result.dividends.append(DividendEntry(
                    source="zerodha",
                    scrip_name=symbol,
                    isin=isin,
                    ex_date=ex_date,
                    amount=amount,
                    tds=0.0,
                ))

    wb.close()
    return result
