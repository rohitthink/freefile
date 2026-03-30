from typing import Optional
import pdfplumber
from pathlib import Path

from .base import BankStatementParser
from .hdfc import HDFCParser
from .sbi import SBIParser
from .icici import ICICIParser
from .axis import AxisParser
from .kotak import KotakParser
from .csv_generic import CSVGenericParser
from .sbi_xlsx import SBIXlsxParser
from .icici_xls import ICICIXlsParser

ALL_PDF_PARSERS: list[BankStatementParser] = [
    ICICIParser(),   # Must be before HDFC/SBI (their keyword fallbacks are too greedy)
    AxisParser(),
    KotakParser(),
    HDFCParser(),    # Keyword fallback — keep last
    SBIParser(),     # Keyword fallback — keep last
]

ALL_XLSX_PARSERS: list[BankStatementParser] = [
    SBIXlsxParser(),
]

ALL_XLS_PARSERS: list[BankStatementParser] = [
    ICICIXlsParser(),
]


def detect_parser(file_path: str, password: Optional[str] = None) -> BankStatementParser:
    """Auto-detect which bank parser to use based on file content."""
    path = Path(file_path)

    if path.suffix.lower() in (".csv", ".tsv"):
        return CSVGenericParser()

    # XLSX detection (e.g. SBI encrypted statements)
    if path.suffix.lower() == ".xlsx":
        return _detect_xlsx(file_path, password)

    # XLS detection (e.g. ICICI OLE2 statements)
    if path.suffix.lower() == ".xls":
        return _detect_xls(file_path)

    # PDF detection
    try:
        open_kwargs = {}
        if password:
            open_kwargs["password"] = password

        with pdfplumber.open(file_path, **open_kwargs) as pdf:
            # Extract text from first 2 pages for detection
            text = ""
            tables = []
            for page in pdf.pages[:2]:
                page_text = page.extract_text() or ""
                text += page_text + "\n"
                tables.extend(page.extract_tables() or [])

            for parser in ALL_PDF_PARSERS:
                if parser.can_parse(text, tables):
                    return parser

    except Exception:
        pass

    # Fallback: try HDFC parser (most generic tabular format)
    return HDFCParser()


def _detect_xlsx(file_path: str, password: Optional[str] = None) -> BankStatementParser:
    """Detect the right parser for .xlsx files by peeking at content."""
    try:
        import io
        import openpyxl

        # Try to open (decrypt if needed)
        if password:
            import msoffcrypto
            with open(file_path, "rb") as f:
                office_file = msoffcrypto.OfficeFile(f)
                office_file.load_key(password=password)
                decrypted = io.BytesIO()
                office_file.decrypt(decrypted)
                decrypted.seek(0)
                wb = openpyxl.load_workbook(decrypted, data_only=True)
        else:
            wb = openpyxl.load_workbook(file_path, data_only=True)

        ws = wb.active
        if ws is None:
            return SBIXlsxParser()  # fallback

        # Collect text from first 20 rows for detection
        text_parts = []
        for row in ws.iter_rows(max_row=20, values_only=True):
            for cell in row:
                if cell is not None:
                    text_parts.append(str(cell))
        text = " ".join(text_parts)

        for parser in ALL_XLSX_PARSERS:
            if parser.can_parse(text, []):
                return parser

    except Exception:
        pass

    # Fallback to SBI XLSX parser (most common encrypted xlsx)
    return SBIXlsxParser()


def _detect_xls(file_path: str) -> BankStatementParser:
    """Detect the right parser for .xls (OLE2) files by peeking at content."""
    try:
        import xlrd
        wb = xlrd.open_workbook(file_path)
        ws = wb.sheet_by_index(0)

        # Collect text from first 20 rows for detection
        text_parts = []
        for i in range(min(20, ws.nrows)):
            for j in range(ws.ncols):
                val = ws.cell_value(i, j)
                if val:
                    text_parts.append(str(val))
        text = " ".join(text_parts)

        for parser in ALL_XLS_PARSERS:
            if parser.can_parse(text, []):
                return parser

    except Exception:
        pass

    # Fallback to ICICI XLS parser (most common xls format)
    return ICICIXlsParser()
