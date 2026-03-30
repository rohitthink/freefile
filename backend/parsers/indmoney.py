"""Parser for INDmoney Tax Report XLSX files.

Handles two types:
1. US Stocks Tax Report (tax_us-stocks_*.xlsx) — trades, dividends, Schedule FSI
2. EQ/F&O Tax P&L (Indmoney-AnnualTaxPnlReport-*.xlsx) — domestic trades
"""

import openpyxl
from dataclasses import dataclass, field
from datetime import datetime
from typing import Optional


@dataclass
class USStockEntry:
    """A single US stock trade entry."""
    source: str = "indmoney"
    scrip_name: str = ""
    isin: str = ""
    asset_type: str = "us_stock"
    buy_date: Optional[str] = None
    sell_date: Optional[str] = None
    quantity: float = 0.0
    buy_value: float = 0.0  # in INR
    sell_value: float = 0.0  # in INR
    buy_value_usd: float = 0.0
    sell_value_usd: float = 0.0
    expenses: float = 0.0  # in INR
    expenses_usd: float = 0.0
    exchange_rate: float = 0.0
    gain_loss: float = 0.0  # in INR
    gain_type: str = ""
    holding_period_days: int = 0
    broker: str = ""


@dataclass
class USDividendEntry:
    """A US stock dividend entry."""
    source: str = "indmoney"
    scrip_name: str = ""
    date: Optional[str] = None
    amount_usd: float = 0.0
    amount_inr: float = 0.0
    exchange_rate: float = 0.0
    tds_usd: float = 0.0
    tds_inr: float = 0.0


@dataclass
class ScheduleFSIEntry:
    """Schedule FSI (Foreign Source Income) entry for ITR."""
    country: str = "USA"
    income_type: str = ""
    income_inr: float = 0.0
    tax_paid_outside: float = 0.0
    tax_relief_available: float = 0.0


@dataclass
class INDmoneyParseResult:
    """Complete result from parsing INDmoney report."""
    trades: list = field(default_factory=list)
    dividends: list = field(default_factory=list)
    schedule_fsi: list = field(default_factory=list)
    summary: dict = field(default_factory=dict)


def _sf(val) -> float:
    """Safe float conversion."""
    if val is None:
        return 0.0
    try:
        f = float(val)
        return round(f, 2) if f == f else 0.0
    except (ValueError, TypeError):
        if isinstance(val, str):
            cleaned = val.strip().replace(",", "").replace("$", "").replace("₹", "").strip()
            if cleaned and cleaned != "-":
                try:
                    return round(float(cleaned), 2)
                except (ValueError, TypeError):
                    pass
        return 0.0


def _ss(val) -> str:
    """Safe string conversion."""
    if val is None:
        return ""
    return str(val).strip()


def _pd(val) -> Optional[str]:
    """Parse date to ISO format."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    if not s or s == "-":
        return None
    for fmt in ["%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y", "%d-%b-%Y"]:
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    try:
        from dateutil import parser as dp
        return dp.parse(s, dayfirst=True).strftime("%Y-%m-%d")
    except (ValueError, TypeError):
        return None


def parse_indmoney_us(file_path: str) -> INDmoneyParseResult:
    """Parse INDmoney US Stocks Tax Report XLSX."""
    result = INDmoneyParseResult()
    wb = openpyxl.load_workbook(file_path, data_only=True)

    # Detect file type
    sheet_names_lower = {s.lower(): s for s in wb.sheetnames}

    # Check if this is a US stocks file or domestic EQ file
    is_us_stocks = any("us stock" in s for s in sheet_names_lower)
    is_eq_tradewise = any("tradewise" in s for s in sheet_names_lower)

    if is_us_stocks:
        _parse_us_stocks_file(wb, sheet_names_lower, result)
    elif is_eq_tradewise:
        _parse_eq_file(wb, sheet_names_lower, result)
    else:
        # Try first sheet as US stocks
        _parse_us_stocks_file(wb, sheet_names_lower, result)

    wb.close()
    return result


def _parse_us_stocks_file(wb, sheet_map: dict, result: INDmoneyParseResult):
    """Parse INDmoney US Stocks format."""
    # Find US Stocks sheet
    us_sheet = None
    for k, v in sheet_map.items():
        if "us stock" in k:
            us_sheet = v
            break
    if not us_sheet:
        us_sheet = wb.sheetnames[0] if wb.sheetnames else None
    if not us_sheet:
        return

    ws = wb[us_sheet]

    # Extract summary from top rows
    for r in range(1, min(20, ws.max_row + 1)):
        label = _ss(ws.cell(r, 1).value).lower()
        if "- short-term" in label:
            result.summary["stcg_inr"] = _sf(ws.cell(r, 2).value)
        elif "- long-term" in label:
            result.summary["ltcg_inr"] = _sf(ws.cell(r, 2).value)
        elif "- dividend" in label:
            result.summary["dividends_inr"] = _sf(ws.cell(r, 2).value)
        elif "- interest" in label:
            result.summary["interest_inr"] = _sf(ws.cell(r, 2).value)

    # Parse trade sections
    # Structure: section header row, then 2-row header, then "Gains"/"Losses" sub-headers, then data
    # Columns: A=Name, B=Sell date, C=Purchase date, D=Qty, E=Sell $/unit, F=Buy $/unit,
    #          G=Sell $ total, H=Buy $ total, I=Sell expense $, J=Buy expense $,
    #          K=Sell FX, L=Buy FX, M=Sell INR, N=Buy INR, O=Sell exp INR, P=Buy exp INR,
    #          Q=Taxable Gains, R=Broker
    current_section = None
    for r in range(1, ws.max_row + 1):
        cell_a = _ss(ws.cell(r, 1).value)
        cell_a_lower = cell_a.lower()

        # Detect section headers
        if "short-term" in cell_a_lower and "capital" in cell_a_lower:
            current_section = "short_term"
            continue
        elif "long-term" in cell_a_lower and "capital" in cell_a_lower:
            current_section = "long_term"
            continue
        elif "dividend" in cell_a_lower and "categorised" in cell_a_lower:
            current_section = "dividend"
            continue
        elif "interest" in cell_a_lower and "categorised" in cell_a_lower:
            current_section = "interest"
            continue
        elif "disclaimer" in cell_a_lower:
            break

        if current_section in ("short_term", "long_term"):
            # Skip header rows, "Gains"/"Losses" labels
            if cell_a_lower in ("name of stock", "gains", "losses", "", "total") or \
               cell_a_lower.startswith("sell value") or cell_a_lower.startswith("purchase"):
                if cell_a_lower == "total":
                    continue
                continue
            if not cell_a or cell_a == "-":
                continue

            # This is a trade row
            name = cell_a
            sell_date = _pd(ws.cell(r, 2).value)
            buy_date = _pd(ws.cell(r, 3).value)
            qty = _sf(ws.cell(r, 4).value)
            sell_usd = _sf(ws.cell(r, 7).value)
            buy_usd = _sf(ws.cell(r, 8).value)
            sell_exp_usd = _sf(ws.cell(r, 9).value)
            buy_exp_usd = _sf(ws.cell(r, 10).value)
            sell_fx = _sf(ws.cell(r, 11).value)
            buy_fx = _sf(ws.cell(r, 12).value)
            sell_inr = _sf(ws.cell(r, 13).value)
            buy_inr = _sf(ws.cell(r, 14).value)
            sell_exp_inr = _sf(ws.cell(r, 15).value)
            buy_exp_inr = _sf(ws.cell(r, 16).value)
            taxable = _sf(ws.cell(r, 17).value)
            broker = _ss(ws.cell(r, 18).value)

            holding_days = 0
            if buy_date and sell_date:
                try:
                    bd = datetime.strptime(buy_date, "%Y-%m-%d")
                    sd = datetime.strptime(sell_date, "%Y-%m-%d")
                    holding_days = (sd - bd).days
                except (ValueError, TypeError):
                    pass

            # US stocks: STCG if <=24 months, LTCG Section 112 if >24 months
            gain_type = "STCG_slab" if current_section == "short_term" else "LTCG_112"

            result.trades.append(USStockEntry(
                scrip_name=name,
                buy_date=buy_date,
                sell_date=sell_date,
                quantity=qty,
                buy_value=buy_inr,
                sell_value=sell_inr,
                buy_value_usd=buy_usd,
                sell_value_usd=sell_usd,
                expenses=sell_exp_inr + buy_exp_inr,
                expenses_usd=sell_exp_usd + buy_exp_usd,
                exchange_rate=sell_fx,
                gain_loss=taxable,
                gain_type=gain_type,
                holding_period_days=holding_days,
                broker=broker,
            ))

        elif current_section == "dividend":
            # Dividend: A=Name, B=Date, C=FX Rate, D=USD amt, E=INR amt, F=TDS USD, G=TDS INR, H=Broker
            if cell_a_lower in ("name of the company", "total", "") or not cell_a or cell_a == "-":
                continue

            result.dividends.append(USDividendEntry(
                scrip_name=cell_a,
                date=_pd(ws.cell(r, 2).value),
                amount_usd=_sf(ws.cell(r, 4).value),
                amount_inr=_sf(ws.cell(r, 5).value),
                exchange_rate=_sf(ws.cell(r, 3).value),
                tds_usd=_sf(ws.cell(r, 6).value),
                tds_inr=_sf(ws.cell(r, 7).value),
            ))

    # Parse Schedule FSI
    fsi_sheet = None
    for k, v in sheet_map.items():
        if "fsi" in k:
            fsi_sheet = v
            break

    if fsi_sheet:
        ws = wb[fsi_sheet]
        for r in range(1, ws.max_row + 1):
            # Look for rows with income types
            for c in range(1, ws.max_column + 1):
                cell_val = _ss(ws.cell(r, c).value).lower()
                if cell_val in ("capital gains", "capital gain"):
                    income = _sf(ws.cell(r, c + 1).value)
                    tax_paid = _sf(ws.cell(r, c + 2).value)
                    if income:
                        result.schedule_fsi.append(ScheduleFSIEntry(
                            income_type="capital_gains",
                            income_inr=income,
                            tax_paid_outside=tax_paid,
                        ))
                elif cell_val == "dividend":
                    income = _sf(ws.cell(r, c + 1).value)
                    tax_paid = _sf(ws.cell(r, c + 2).value)
                    if income:
                        result.schedule_fsi.append(ScheduleFSIEntry(
                            income_type="dividend",
                            income_inr=income,
                            tax_paid_outside=tax_paid,
                        ))

    # Build summary
    result.summary["total_stcg_inr"] = round(sum(t.gain_loss for t in result.trades if t.gain_type == "STCG_slab"), 2)
    result.summary["total_ltcg_inr"] = round(sum(t.gain_loss for t in result.trades if t.gain_type == "LTCG_112"), 2)
    result.summary["total_dividends_inr"] = round(sum(d.amount_inr for d in result.dividends), 2)
    result.summary["total_tds_inr"] = round(sum(d.tds_inr for d in result.dividends), 2)
    result.summary["total_trades"] = len(result.trades)


def _parse_eq_file(wb, sheet_map: dict, result: INDmoneyParseResult):
    """Parse INDmoney domestic EQ/F&O Tax P&L format."""
    # Find tradewise sheet
    tradewise_sheet = None
    summary_sheet = None
    fno_sheet = None

    for k, v in sheet_map.items():
        if "tradewise" in k:
            tradewise_sheet = v
        elif "tax p&l" in k and "f&o" not in k:
            summary_sheet = v
        elif "f&o" in k:
            fno_sheet = v

    # Parse summary
    if summary_sheet:
        ws = wb[summary_sheet]
        for r in range(1, min(20, ws.max_row + 1)):
            label = _ss(ws.cell(r, 1).value).lower()
            val = _sf(ws.cell(r, 2).value)
            if "intraday" in label:
                result.summary["intraday_pnl"] = val
            elif "short-term" in label:
                result.summary["short_term_pnl"] = val
            elif "long-term" in label:
                result.summary["long_term_pnl"] = val

    # Parse tradewise exits
    if tradewise_sheet:
        ws = wb[tradewise_sheet]
        current_section = None

        for r in range(1, ws.max_row + 1):
            cell_a = _ss(ws.cell(r, 1).value)
            cell_a_lower = cell_a.lower()

            # Section headers
            if "intraday" in cell_a_lower and "exits" in cell_a_lower:
                current_section = "intraday"
                continue
            elif "short-term" in cell_a_lower and "exits" in cell_a_lower:
                current_section = "short_term"
                continue
            elif "long-term" in cell_a_lower and "exits" in cell_a_lower:
                current_section = "long_term"
                continue

            if not current_section:
                continue

            # Skip headers
            if cell_a_lower in ("stock name", "total", "") or not cell_a:
                continue

            # Data row: A=Stock Name, B=ISIN, C=Buy Date, D=Sell Date, E=Qty,
            #           F=Buy Amount, G=Sell Amount, H=Profit/Loss, I=Holding Days
            isin = _ss(ws.cell(r, 2).value)
            buy_date = _pd(ws.cell(r, 3).value)
            sell_date = _pd(ws.cell(r, 4).value)
            qty = _sf(ws.cell(r, 5).value)
            buy_val = _sf(ws.cell(r, 6).value)
            sell_val = _sf(ws.cell(r, 7).value)
            pnl = _sf(ws.cell(r, 8).value)
            holding = int(_sf(ws.cell(r, 9).value))

            if qty == 0 and buy_val == 0:
                continue

            gain_type_map = {
                "intraday": "speculative",
                "short_term": "STCG_111A",
                "long_term": "LTCG_112A",
            }

            result.trades.append(USStockEntry(
                source="indmoney",
                scrip_name=cell_a,
                isin=isin,
                asset_type="equity" if current_section != "intraday" else "equity_intraday",
                buy_date=buy_date,
                sell_date=sell_date,
                quantity=qty,
                buy_value=buy_val,
                sell_value=sell_val,
                gain_loss=pnl,
                gain_type=gain_type_map.get(current_section, "STCG_111A"),
                holding_period_days=holding,
            ))

    # Parse F&O summary
    if fno_sheet:
        ws = wb[fno_sheet]
        for r in range(1, min(20, ws.max_row + 1)):
            label = _ss(ws.cell(r, 1).value).lower()
            val = _sf(ws.cell(r, 2).value)
            if "total f&o" in label and "p&l" in label:
                result.summary["fno_pnl"] = val

    result.summary["total_trades"] = len(result.trades)
